'''

    操作EXcel表格
'''
import copy

import openpyxl
from openpyxl.styles import Font,colors,Border,Side
import time
import calendar
import string
import os
import json
import shutil
import pprint
import re

import abcd
# 每一个月多少天
MON =  calendar.mdays
# A-Z的大写字母
# AZ = string.ascii_uppercase
AZ = abcd.letterAable()


class MyExcel:
    def __init__(self,dylist:list):
        '''
        
            self._path: 路径
            self._number: 抖音号个数
            self._twoTitleList: Excel第二栏数据(按钮列表)
            self._twoFile: 存放动态添加按钮的数据文件
            self._twoSignalDict: Excel第二栏数据信号字典(对应每个按钮的信号)
            self._gridNumber: 格子数量
            self._DY: 当前操作的抖音号
            self._DYList: 抖音号列表
            self._ExcelName: 自动生成Excel的名称(无后缀)
            self._year, self._mon, self._day: 年,月,日
            self._yearMon: 年_月
            self._ExcelName_: 自动生成Excel的名称(有后缀)
            self._path_Name: 路径+名称
            self._title: Excel标题
            self._xuhao: 序号(Excel中的序号,EG: A,B,C...)
            self._oldxuhao: 旧序号(保存序号修改前的序号)
            self._excelDataDict: Excel文件数据字典
            self._btnName: 设置当前操作按钮的名字
            self._dataing: 当前修改的数据
            self._oldExtime: 已经创建过的Excel
            self._oldExtimeFileName: 存放已经创建过的Excel数据文件
        '''
        self._path = os.getcwd()
        # 抖音号个数
        if len(dylist) <= 1:
            self._number = 1
        else:
            self._number = len(dylist)
            
        self._twoFile = "two.json"
        self._twoTitleList = ["打招呼次数", "回复次数"]
        try:
            with open(self._path + "/" + self._twoFile,"r") as f:
                self._twoTitleList = json.load(f)
        except Exception:
            with open(self._path+"/"+self._twoFile,"w") as f:
                json.dump(self._twoTitleList,f)
        self._twotileLen = len(self._twoTitleList) # 长度

        self._twoSignalDict = {k:False for k in self._twoTitleList}
        self._gridNumber = self._number * self._twotileLen + 1
        self._DY = "456"

        if dylist:
            self._DYList = dylist
        else:
            self._DYList = ["demo"]

        self._ExcelName = time.strftime("%Y_%m_%d", time.localtime())
        self._year, self._mon, self._day = self._ExcelName.split("_")
        self._yearMon = self._year + "_" +self._mon
        self._ExcelName_ = self._yearMon + ".xlsx"
        self._path_Name = self._path+"/"+self._ExcelName_
        self._title = "抖音招呼记录"
        self._xuhao = list(AZ[1:self._gridNumber])
        self._oldxuhao = {"oldx": self._xuhao, "oldLen": len(self._twoTitleList), "type": True, "k": None}  # 第二栏旧数据
        self._excelDataDict = dict()
        self._btnName = ""
        self._dataing = None
        self._oldExtime = []
        self._oldExtimeFileName = self._path+"/"+"oldExtime.json"

        if not os.path.exists(self._oldExtimeFileName):
            with open(self._oldExtimeFileName, "w") as f:
                json.dump([],f)
        else:
            with open(self._oldExtimeFileName, "r") as f:
                 self._oldExtime = json.load(f)

        # 一个月自动创建一个EXcel表格
        try:
            if self._yearMon not in self._oldExtime:
                self.crateExcel()
                # 加入历史
                self._oldExtime.append(self._yearMon)
                with open(self._oldExtimeFileName, "w") as f:
                    json.dump(self._oldExtime,f)
        except Exception:
            # 创建一个工作簿
            pass

    # 居中
    def _myAlignment(self):
        return openpyxl.styles.Alignment(horizontal="center", vertical="center")

    # 边框
    def _myBorder(self):
       return Border(top=Side(style='medium', color='FF000000'),
               left=Side(style='medium', color='FF000000'),
               right=Side(style='medium', color='FF000000'),
               bottom=Side(style='medium', color='FF000000'))
    # 字体
    def _myFont(self):
        return Font(name='等线', size=11, italic=False, color=colors.BLACK, bold=True)

    # 给外部调用,设置那个按钮为True
    def setTwoSignalDict(self,k):
        self._twoSignalDict[k] = True

    # 设置当前操作按钮的名字
    def setBtnName(self,name:str):
        self._btnName = name
    # 获取当前操作按钮的名字
    def getBtnName(self):
        return self._btnName
    # 当前当数据
    def getData(self):
        return self._dataing

    # 给外部设置抖音号
    def setDY(self,number:str):
        self._DY = number

    # 当前操作的抖音号
    def dy(self):
        return self._DY

    # 序号和与DyList同步
    def _xuhaoDySyn(self):
        self._twotileLen = len(self._twoTitleList)  # 长度
        # 第二栏数据信号字典(对应每个按钮的信号)
        self._twoSignalDict = {k: False for k in self._twoTitleList}
        # 格子数量
        self._gridNumber = len(self.getDyList()) * self._twotileLen + 1
        self._xuhao = list(AZ[1:self._gridNumber])

    # 旧数据同步
    def oldDataSyn(self):
        self._oldxuhao = {"oldx": self._xuhao.copy(), "oldLen": len(self._twoTitleList), "type": True, "k": None}

    # 给外部设置抖音号列表
    def setDYList(self,nlist):
        self._DYList = nlist
        self._xuhaoDySyn()  # 同步

    # 增加抖音列表
    def addDYList(self,text):
        self._DYList.append(text)
        self._xuhaoDySyn()  # 同步

    # 删除抖音列表
    def delDYList(self,text):
        self._DYList.remove(text)
        self._xuhaoDySyn()  # 同步

    # 获取抖音列表
    def getDyList(self):
        return self._DYList

    # 增加第二栏第数据
    def addTwoList(self,k):
        # 恢复
        self._oldxuhao = {"oldx":self._xuhao.copy(),"oldLen":len(self._twoTitleList),"type":True,"k":k} # 旧数量
        self._twoTitleList.append(k)
        # 写入本地
        with open(self._path+"/"+self._twoFile,"w") as f:
            json.dump(self._twoTitleList,f)
        self._xuhaoDySyn()  # 同步

    # 删除第二栏第数据
    def delTwoList(self,k):
        # 恢复
        self._oldxuhao = {"oldx": self._xuhao.copy(), "oldLen": len(self._twoTitleList),"type":False,"k":k}  # 旧数量
        self._twoTitleList.remove(k)
        # 写入本地
        with open(self._path + "/" + self._twoFile,"w") as f:
            json.dump(self._twoTitleList, f)
        self._xuhaoDySyn()  # 同步

    # 获取第二栏的数据
    def getTwoList(self):
        return self._twoTitleList

    # 当前月的天数
    def mon(self):
        return MON[int(self._mon)]

    # 创建空EXcel
    def crateExcel(self):
        _xuhao = self._xuhao.copy()
        _xuhao.insert(0,"A")

        # 创建单元
        self._ex = openpyxl.Workbook()
        sheet = self._ex.active
        #
        for i in _xuhao:
            for j in range(1, MON[int(self._mon)] + +4):
                sheet["{}".format(i+str(j))].border = self._myBorder()

        # 创建标题
        sheet.column_dimensions["A"].width = 10
        for i in self._xuhao:
            sheet.column_dimensions[i].width = 13
        sheet.row_dimensions[1].height = 40  # 标题高度
        sheet.merge_cells("{}:{}".format("A1", self._xuhao[-1] + "1"))   # 合并
        sheet["A1"] = self._title   # 设置标题
        sheet["A1"].alignment = self._myAlignment()  # 设置居中
        # 创建第二栏
        x = 0
        
        for i in range(0,len(self._xuhao)-1,self._twotileLen):
            sheet.merge_cells("{}:{}".format(self._xuhao[i]+"2",self._xuhao[i+(self._twotileLen-1)]+"2"))
            sheet[self._xuhao[i] + "2"] = "抖音号:"+self._DYList[x]  # 设置标题
            sheet[self._xuhao[i] + "2"].alignment = self._myAlignment() # 设置居中
            x+=1
            
        # 第三栏
        sheet["A3"] = "日期"
        sheet["A3"].alignment = self._myAlignment()  # 设置居中
        index = 0
        maxIndex = len(self._twoTitleList)
        for i in self._xuhao:
            if index < maxIndex:
                sheet[i + "3"] = self._twoTitleList[index]
                index += 1
                if index == maxIndex:
                    index = 0
            sheet[i + "3"].alignment = self._myAlignment()  # 设置居中
        # 第四栏
        for i in range(1,MON[int(self._mon)]+1):
            sheet["{}".format('A'+str(i+3))] = i
            sheet["{}".format('A'+str(i+3))].alignment = self._myAlignment()  # 设置居中

        self._ex.save(self._path_Name)
        self._ex.close()

    # 获取Excel  x,y的值
    def get_xy_value(self,x,y):
        self._openExcel = openpyxl.load_workbook(self._path_Name)
        sh = self._openExcel[self._openExcel.sheetnames[0]]

        v = sh.cell(row=x, column=y)
        tempv = v.value
        # self._openExcel.save(self._path_Name)
        self._openExcel.close()
        return tempv

    # 设置Excel  x,y的值
    def set_xy_value(self,x,y,value):
        self._openExcel = openpyxl.load_workbook(self._path_Name)
        sh = self._openExcel[self._openExcel.sheetnames[0]]

        sh.cell(row=x, column=y, value=value).alignment=self._myAlignment()

        self._openExcel.save(self._path_Name)
        self._openExcel.close()
        # return tempv

    # Excel名称
    def getEName(self):
        return self._ExcelName_
    # 返回路径
    def getEPath(self):
        return self._path

    # 返回Excel数据字典
    def ExcelDict(self):
        return self._excelDataDict

    # 抖音和单元序号做映射
    def _Dytoxuhao(self):
        temp =dict()
        x = 0
        # 格子数量
        self._gridNumber = len(self.getDyList()) * len(self._twoTitleList) + 1
        self._xuhao = list(AZ[1:self._gridNumber])
        # print("xuhao:",self._xuhao," len:",len(self._xuhao))
        # print("twolist:",len(self._twoTitleList))
        for dy in self._DYList:
            temp[dy] = dict()
            x_ = x
            # print("_x:",x_)
            for k in self._twoTitleList:
                temp[dy][k] = self._xuhao[x_]
                x_ += 1
            x += len(self._twoTitleList)

        return temp

    # 抖音和单元序号做反向映射
    def _xuhaotoDy(self,unit:str):
        temp = self._Dytoxuhao()
        for k,v in temp.items():
            for zk in self._twoTitleList:
                if unit == temp[k][zk]:
                    return (k,zk,v)

    # 解析Excel字典(需要先读取readExcel)
    def parsingExcelDict(self):
        # temp = self._Dytoxuhao()

        parEx = dict()
        # 先读取readExcel
        self.readExcel()
        for ek,kv in self.ExcelDict().items():
            listk = re.findall("[A-Z].*?", ek)
            head = "".join(listk)  # 取单元号头
            listek = re.findall("[0-9].*?",ek)
            # rq = ek[1:]  # 第几天
            rq = "".join(listek)  # 第几天
            dy = self._xuhaotoDy(head)
            if dy:  # 防止在删除数据时,引发的报错
                dyh = dy[0]  # 抖音号
                if dyh not in parEx:
                    parEx[dyh] = {rq:{dy[1]:kv}}
                else:
                    if rq not in parEx[dyh]:
                        parEx[dyh][rq] = {dy[1]: kv}
                    else:
                        parEx[dyh][rq][dy[1]] = kv

        return parEx

    # 统计 打招呼 和 回复 的次数
    def helloReplyData(self):
        temp = self.parsingExcelDict()

        result = dict()
        for fk, fv in temp.items():
            for zv in fv.values():
                for k,v in zv.items():
                    if fk not in result:
                        result[fk] = {k: 0 for k in self._twoTitleList}
                        if k in self._twoTitleList:
                            if type(v) == int:
                                result[fk][k] += v
                            elif type(v) == str:
                                result[fk][k] = v
                    else:
                        if type(v) == int:
                            result[fk][k] += v
                        elif type(v) == str:
                            result[fk][k] = v

        return result


    # 增加
    def AExe(self,data:str="",AR:bool=True,ExcelName=None):
        if not ExcelName:
            self._openExcel= openpyxl.load_workbook(self._path_Name)
            sh = self._openExcel[self._openExcel.sheetnames[0]]

            # 获取第二行对象(除去了日期)
            san=sh[2][1:-1]
            temp = []   # 存储抖音号单元格的对象
            DYnumnerData = []  # 存储抖音号
            DYtoUnit = dict()  # 抖音对应单元格
            for i in range(0,len(san),self._twotileLen):
                temp.append(san[i])
            for i in temp:
                DYnumnerData.append(i.value)
            i = 0


            for name in DYnumnerData:
                name = name.split(":")[-1]
                DYtoUnit[name]=list()
                index = i
                for k in self._twoTitleList:
                    DYtoUnit[name].append({k:self._xuhao[index]})
                    index += 1

                i += self._twotileLen

            if self._DY in DYtoUnit:
                col: str
                ClickName = ""  #保存被点击的按钮的名字
                for k,v in self._twoSignalDict.items():
                    if v:
                        ClickName = k
                col = DYtoUnit[self._DY][self._twoTitleList.index(ClickName)][ClickName]
                self._twoSignalDict[ClickName] = False

                row = int(self._day) + 3
                row = str(row)
                v = sh[col + row].value

                datav=data
                if not v:
                    if data:
                        sh[col + row] = datav
                    elif AR:
                        datav = 1
                        sh[col + row] = datav
                    else:
                        sh[col + row] = datav
                else:
                    if data:
                        sh[col + row] = datav
                    elif AR:
                        try: # 读取数据时的问题
                            datav = int(v) + 1
                            sh[col + row] = datav
                        except Exception as e:
                            # print("错误009:",e)
                            datav = 1
                            sh[col + row] = datav
                    else:
                        datav = int(v) - 1
                        if datav <0:
                            datav=0
                        if datav:
                            sh[col + row] = datav
                        else:
                            datav = 0
                            sh[col + row] = datav
                self._dataing = datav
                sh[col + row].alignment = self._myAlignment()  # 设置居中

            self._openExcel.save(self._path_Name)
            self._openExcel.close()

    # 下载Excel
    def downExcel(self,newFile):
        shutil.copy(self._path_Name,newFile)

    # 获取一个小标题所有位置
    def getStitlePos(self,title):
        self._openExcel = openpyxl.load_workbook(self._path_Name)
        sh = self._openExcel[self._openExcel.sheetnames[0]]
        row = MON[int(self._mon)] + 2

        temp = []
        traverse = self._xuhao
        # 当前者长度大于后者时,代表删除操作
        if len(self._oldxuhao["oldx"])>len(self._xuhao):
            traverse = self._oldxuhao["oldx"]

        for i in traverse:
            if title == sh[i+str(3)].value:
                temp.append(i+"3")
        # self._openExcel.save(self._path_Name)
        self._openExcel.close()
        return temp

    # 清除一个小标题的所有数据
    def clearKey(self,title):
        # 获取位置
        pos = self.getStitlePos(title)

        self._openExcel = openpyxl.load_workbook(self._path_Name)
        sh = self._openExcel[self._openExcel.sheetnames[0]]
        row = MON[int(self._mon)] + 3
        for k in pos:
            head = k[0]
            for i in range(4, row + 1):
                sh[head+str(i)]=None

        self._openExcel.save(self._path_Name)
        self._openExcel.close()

    # 读取Excel数据
    def readExcel(self):
        # 格子数量
        # self._gridNumber = len(self.getDyList()) * len(self._twoTitleList) + 1
        # self._xuhao = list(AZ[1:self._gridNumber])

        self._openExcel = openpyxl.load_workbook(self._path_Name)
        sh = self._openExcel[self._openExcel.sheetnames[0]]
        row = MON[int(self._mon)] + 3

        if not self._oldxuhao["k"]:
            self.oldDataSyn() # 就数据同步
        # 清空数据
        self._excelDataDict.clear()

        # 旧数据分组
        oldxuhao = []
        oldtwoLen = self._oldxuhao["oldLen"]
        s, e = 0, oldtwoLen
        for i in range(len(self._DYList)):
            oldxuhao.append(self._oldxuhao["oldx"][s:e])
            s = e
            e += oldtwoLen

        # 新数据分组
        newXuhao = []
        twoLen = len(self._twoTitleList)
        s, e = 0, twoLen

        for i in range(len(self._DYList)):
            newXuhao.append(self._xuhao[s:e])
            s = e
            e += twoLen

        # 当前者长度大于后者时,代表删除操作
        if len(oldxuhao[0])>len(newXuhao[0]):
            # 获取位置,只有字母部分
            temp = [v[0] for v in self.getStitlePos(self._oldxuhao["k"])]
            oldLen = len(oldxuhao) #长度
            for v in temp:
                for i in range(oldLen):
                    if v in oldxuhao[i]:
                        oldxuhao[i].remove(v)
        # 读取数据
        for ex in range(len(oldxuhao)):
            for oldi in range(len(oldxuhao[ex])):
                for i in range(4, row + 1):
                    oldx = oldxuhao[ex][oldi]
                    v = sh["{}".format(oldx + str(i))].value
                    if v and oldi < len(newXuhao[0]): # 存有效数据,随便防止越界
                        newx = newXuhao[ex][oldi]
                        self._excelDataDict[newx + str(i)] = v

        # pprint.pprint(self._excelDataDict)
        self._openExcel.save(self._path_Name)
        self._openExcel.close()

    # 写数据
    def writeExcel(self,path=None):
        if not path:
            path=self._path_Name
        self._openExcel = openpyxl.load_workbook(path)
        sh = self._openExcel[self._openExcel.sheetnames[0]]
        # 需要删除的数据列表
        delList = []

        for k,v in self._excelDataDict.items():
            listk = re.findall("[A-Z].*?",k)
            head = "".join(listk)
            # 防止写入多余数据
            if head in self._xuhao:
                sh[k]=v
                sh[k].alignment = self._myAlignment() # 居中
            else:
                delList.append(k)
        # 删除
        if delList:
            for i in delList:
                del self._excelDataDict[i]

        self._openExcel.save(path)
        self._openExcel.close()

    def write(self):
        self._openExcel = openpyxl.load_workbook(self._path_Name)
        sh = self._openExcel[self._openExcel.sheetnames[0]]

        for row in self._xuhao:
            for col in range(4,self.mon()+4):
                sh[row+str(col)] = 1


        self._openExcel.save(self._path_Name)
        self._openExcel.close()


    # 扩展第而栏第大小,需要先addTwoList("奥特曼")
    def extensionTwo(self):
        self.readExcel()
        self.crateExcel()
        self.writeExcel()

    # 设置Excel大小，需要先添加addDYList("奥")
    def extensionExcel(self,dynumber:int=-1):
        size = len(self.getDyList())  # 抖音号个数
        if dynumber == -1:
            self._number = size
        elif dynumber >= size:
            self._number = size
        elif dynumber <= size:
            self._number =dynumber

        # 格子数量
        self._gridNumber = len(self.getDyList()) *len(self._twoTitleList) + 1
        self._xuhao = list(AZ[1:self._gridNumber])
        self.readExcel()
        self.crateExcel()
        self.writeExcel()


if __name__ == '__main__':
    m = MyExcel(["AA","BB","CC","DD","EE","FF","HH"])
    # m.write()
    m.readExcel()
    m.writeExcel()
